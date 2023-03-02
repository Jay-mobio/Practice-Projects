from openpyxl import Workbook
import psycopg2
import pandas as pd
# Establish a connection to the database
conn = psycopg2.connect(database="task", user="postgres", password="1713", host="localhost", port="5432")

# Create a cursor object to execute SQL commands
cur = conn.cursor()

# Write the SQL query to join the tables
sql_query = """
SELECT customers.customer_id, orders.order_date, orders.order_total, orderdetails.quantity, orderdetails.price, customers.customer_name, customers.email, customers.phone, customers.address
FROM customers
JOIN orders ON customers.customer_id = orders.customer_id
JOIN orderdetails ON orders.order_id = orderdetails.order_id
"""

# Execute the SQL query
cur.execute(sql_query)

# Fetch all the data returned by the query
results = cur.fetchall()

df = pd.DataFrame(results,columns = ['customer_id', 'order_date', 'order_total', 'quantity', 'price', 'customer_name', 'email', 'phone', 'address'])
df['order_date'] = pd.to_datetime(df['order_date'], format='%Y-%m-%d')

# Create a new DataFrame with the 'order_date' column converted to the desired format
df_new = df.copy()
df_new['order_date'] = df_new['order_date'].dt.strftime('%d-%m-%Y')

wb = Workbook()
ws = wb.active

for col, header in enumerate(df_new.columns):
    ws.cell(row=1, column=col+1, value=header)

for row_idx, row_data in enumerate(df_new.values):
    for col_idx, cell_data in enumerate(row_data):
        if isinstance(cell_data, pd.Timestamp):
            # Set the cell format to "dd-mm-yyyy"
            ws.cell(row=row_idx+2, column=col_idx+1).number_format = numbers.FORMAT_DATE_DDMMYYYY
        ws.cell(row=row_idx+2, column=col_idx+1, value=cell_data)

wb.save("joins.xlsx")
# Close the cursor and connection
cur.close()
conn.close()
